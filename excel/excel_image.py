# @Author  : 记不起的回忆
# @email   : 836875692@qq.com
# @File    : excel_image.py
# @Project : python-modules
# @Software: PyCharm
"""
pip install Pillow
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from pandas import read_excel


class ExcelRead:
    def __init__(self, path, sheet):
        self._path = path
        self.df = read_excel(self._path, sheet_name=sheet, dtype=str)

    def get_columns(self):
        return self.df.columns.values

    def get_rows_by_columns(self, columns):
        """
        读所有行的title以及data列的值，这里需要嵌套列表
        :return:
        """
        data = self.df.loc[:, columns].values
        return data

    def get_rows_by_sheet_columns(self, sheet, columns):
        data = self.df[sheet].loc[:, columns].values
        return data


def open_xls_insert_img(file, img_path):
    """
    插入图片
    :param file Excel文件
    :param img_path: 本地图片路径
    :return:
    """
    wb = load_workbook(file, data_only=True)
    sheet = wb['aa']  # 获取sheet

    # 使用网络图片
    # response = requests.get('网络图片URL')
    # image_bytes = BytesIO()
    # image_bytes.write(response.content)
    # img = Image(image_bytes)

    # 使用本地图片
    img = Image(img_path)

    _from = AnchorMarker(1, 50000, 9, 50000)
    to = AnchorMarker(2, -50000, 10, -50000)
    img.anchor = TwoCellAnchor('twoCell', _from, to)
    sheet.add_image(img)  # 添加图片

    wb.save(file)  # 不要忘记保存


def write_data_excel():
    """
    创建 Excel
    :return:
    """
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title='aa')
    rows = [
        ['now1', 'now2', 'now3', 'now3'],
        [12, 111, 23, 26],
        [11, 43, 55, 13],
        [54, 7672, 333, 433],
        [1, 2, 3, 4],
    ]
    for row in rows:
        ws.append(row)
    wb.save('写入数据.xlsx')


def main():
    try:
        write_data_excel()
        open_xls_insert_img('写入数据.xlsx', '1.jpg')
        print('success')
    except Exception as ex:
        print(ex)


if __name__ == '__main__':
    main()
